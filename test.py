import pandas as pd
from datetime import datetime

def extract_and_create_new_excel(input_file_path, output_file_path=None, sheet_name=None):
    """
    Extract specific columns from an Excel file and create a new Excel file
    
    Parameters:
    input_file_path (str): Path to the input Excel file
    output_file_path (str): Path for the new Excel file (optional)
    sheet_name (str): Name of the sheet to read from (optional)
    
    Returns:
    str: Path of the created Excel file
    """
    
    try:
        # Read the original Excel file
        if sheet_name:
            df = pd.read_excel(input_file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(input_file_path)
        
        # Define the columns to extract
        columns_to_extract = [
            'Resolution [AF KCS Article]',  # Column X
            'Number',                       # Column A
            'Short description',            # Column G
            'Category',                     # Column K
            'Meta',                        # Column M
            'URL',                         # Column O
            'Cause',                       # Column R
            'Issue',                       # Column V
            'Resolution',                  # Column Y
            'Use count',                   # Column AF
            'View count'                   # Column AG
        ]
        
        # Check which columns exist
        existing_columns = [col for col in columns_to_extract if col in df.columns]
        
        if existing_columns:
            # Extract the data
            extracted_data = df[existing_columns].copy()
            
            # Generate output filename if not provided
            if not output_file_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file_path = f"extracted_data_{timestamp}.xlsx"
            
            # Create the new Excel file
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                extracted_data.to_excel(writer, sheet_name='Extracted_Data', index=False)
                
                # Optional: Add a summary sheet
                summary_data = {
                    'Metric': ['Total Rows', 'Total Columns', 'Extraction Date', 'Source File'],
                    'Value': [
                        len(extracted_data),
                        len(extracted_data.columns),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        input_file_path
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            print(f"✅ Successfully created new Excel file: {output_file_path}")
            print(f"📊 Extracted {len(extracted_data)} rows and {len(existing_columns)} columns")
            print(f"📋 Columns included: {', '.join(existing_columns)}")
            
            return output_file_path
            
        else:
            print("❌ No matching columns found in the source file.")
            return None
            
    except FileNotFoundError:
        print(f"❌ Error: Input file '{input_file_path}' not found.")
        return None
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return None

def create_formatted_excel(input_file_path, output_file_path=None):
    """
    Create a new Excel file with formatting and styling
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # First extract the data
    temp_file = extract_and_create_new_excel(input_file_path, "temp_extracted.xlsx")
    
    if temp_file:
        # Generate final output filename
        if not output_file_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file_path = f"formatted_extracted_data_{timestamp}.xlsx"
        
        # Load the temporary file and apply formatting
        wb = load_workbook("temp_extracted.xlsx")
        ws = wb['Extracted_Data']
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Save the formatted file
        wb.save(output_file_path)
        
        # Clean up temporary file
        import os
        os.remove("temp_extracted.xlsx")
        
        print(f"✨ Formatted Excel file created: {output_file_path}")
        return output_file_path

# Main execution
if __name__ == "__main__":
    # Replace with your input Excel file path
    input_excel_file = "your_excel_file.xlsx"
    
    # Option 1: Create a basic new Excel file
    print("Creating basic Excel file...")
    basic_output = extract_and_create_new_excel(
        input_file_path=input_excel_file,
        output_file_path="extracted_columns_basic.xlsx"
    )
    
    # Option 2: Create a formatted Excel file
    print("\nCreating formatted Excel file...")
    formatted_output = create_formatted_excel(
        input_file_path=input_excel_file,
        output_file_path="extracted_columns_formatted.xlsx"
    )
    
    if basic_output:
        print(f"\n📁 Files created successfully!")
        print(f"Basic file: {basic_output}")
        if formatted_output:
            print(f"Formatted file: {formatted_output}")
